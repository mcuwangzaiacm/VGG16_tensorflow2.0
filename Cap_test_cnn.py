import tensorflow as tf
import cv2
import os
import numpy as np
import time
import openpyxl

def main():

    ExcelPath = 'E:\\AOItest\\XLWchatu\\CapData3.xlsx'
    DataName = "Set14_1"
    Start_Col = 1 + 3*3    # 变换数据集
    Start_Row = 2 + 10*2   # 变换方法
    workbook = openpyxl.load_workbook(ExcelPath)  # 新建
    worksheet = workbook.worksheets[6 - 1]
    ##### 加载模型
    model_path = r'E:\AOItest\AOIpyOtherMethods\Cap_VGG\Cap_model_Vgg16_epo50.h5'
    model = tf.keras.models.load_model(model_path)
    # model.summary()  #输出 加载的网络信息
    model.build(input_shape=[None, 64, 64, 3])

    # (a) 从数据集导入
    # (x, y), (x_test, y_test) = Cap_load_data()
    # # x_test = tf.cast(x_test, dtype=tf.float32) / 255.  或者  转为float类型 且归一化
    # x_train = np.array(x/255., dtype=np.float64)
    # y_train = tf.squeeze(y, axis=1) # [n, 1] => [n]
    # x_test = np.array(x_test/255., dtype=np.float64)
    # y_test = tf.squeeze(y_test, axis=1) # [n, 1] => [n]
    # (b) 直接图片读入
    # src = get_img(r"D:\python2017\example\tensorflow2-course-master\test_data\apple3.jpg")

    strat_time = time.time()
    # # res < 0.5 -->  up
    ImageFile1 = r"E:\\AOItest\\XiaoLunWenDataSet\\" + DataName + "\\up\\"
    ImageNames1 = os.listdir(ImageFile1)
    num = 0
    jishu = 0
    for srcName in ImageNames1:
        secc = cv2.imread(ImageFile1+srcName)
        src = get_img(ImageFile1+srcName)
        res = model.predict(src)
        res = round(float(res), 4)
        if res >= 0.5:
            num = num + 1
        jishu = jishu + 1
        if jishu % 1000 == 0:
            print("已检测完", jishu)
    # print("--------------", num)

    ImageFile2 = r"E:\\AOItest\\XiaoLunWenDataSet\\" + DataName + "\\Down\\"
    ImageNames2 = os.listdir(ImageFile2)
    num1 = 0
    for srcName in ImageNames2:
        src = get_img(ImageFile2 + srcName)
        res = model.predict(src)
        res = round(float(res), 4)
        if res < 0.5:
            num1 = num1 + 1
    # print("--------------", num1)

    end_time = time.time()
    alltime = end_time - strat_time
    print(DataName + "已检测完成 耗时：" + str(end_time - strat_time))
    allnum = len(ImageNames1) + len(ImageNames2)
    avgtime = float(alltime / allnum)

    tp, fn, fp, tn = len(ImageNames1) - num, num, num1, len(ImageNames2) - num1
    accuracy = format((tp + tn) / (tp + fn + fp + tn), '.4f')  # 准确率
    # precision = format(tp / (tp + fp), '.4f')  # 精确率
    recall = format(tp / (tp + fn), '.4f')  # 召回率
    F1 = format((tp * 2) / (2 * tp + fn + fp), '.4f')  # F1值（H-mean值）
    MDR = format((fp) / (tp + fn + fp + tn), '.4f')  # F1值（H-mean值）


    worksheet.cell(Start_Row, Start_Col).value = "TP:"
    worksheet.cell(Start_Row+1, Start_Col).value = "FN:"
    worksheet.cell(Start_Row+2, Start_Col).value = "FP:"
    worksheet.cell(Start_Row+3, Start_Col).value = "TN:"

    worksheet.cell(Start_Row+4, Start_Col).value = "准确率:"
    worksheet.cell(Start_Row+5, Start_Col).value = "召回率:"
    worksheet.cell(Start_Row+6, Start_Col).value = "F1值:"
    worksheet.cell(Start_Row+7, Start_Col).value = "MDR:"
    worksheet.cell(Start_Row+8, Start_Col).value = "耗时:"

    # worksheet.cell(Start_Row, Start_Col + 2).value = str(avgtime)
    worksheet.cell(Start_Row, Start_Col + 1).value = str(tp)
    worksheet.cell(Start_Row+1, Start_Col + 1).value = str(fn)
    worksheet.cell(Start_Row+2, Start_Col + 1).value = str(fp)
    worksheet.cell(Start_Row+3, Start_Col + 1).value = str(tn)
    worksheet.cell(Start_Row+4, Start_Col + 1).value = str(accuracy)
    worksheet.cell(Start_Row+5, Start_Col + 1).value = str(recall)
    worksheet.cell(Start_Row+6, Start_Col + 1).value = str(F1)
    worksheet.cell(Start_Row+7, Start_Col + 1).value = str(MDR)
    worksheet.cell(Start_Row+8, Start_Col + 1).value = str(end_time - strat_time)


    print(accuracy, '', recall, '', F1, '', MDR)
    workbook.save(ExcelPath)

    # prob = tf.nn.softmax(res, axis=1)
    # print(prob)
    # pred = tf.argmax(prob, axis=1)
    # pred = tf.cast(pred, dtype=tf.int32)
    # print(int(pred))

def get_img(data_path):
    # Getting image array from path:
    img = cv2.imread(data_path)
    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)     ## 这里 一定要要注意 提出的信息 要和 训练文件格式一样 RGB排列
    img = cv2.resize(img, (64, 64))
    img = img.reshape(1, 64, 64, 3)
    img = np.array(img / 255., dtype=np.float64)
    return img


if __name__ == "__main__":
    main()




